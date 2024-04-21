import openpyxl # pip install openpyxl
import snowflake.connector # pip install snowflake-connector-python
import pandas # pip install pandas
# pip install "snowflake-connector-python[pandas]"

from dataframe_test import write_records_to_excel

from openpyxl.styles import Border, Side , PatternFill

connection = snowflake.connector.connect(
    user='patilhitendra13',
    # authenticator = 'externalbrowser',
    password='Patilhitendra13',
    account='kf94287.ap-south-1.aws',
    warehouse='COMPUTE_WH',
    database='Hiten',
    schema='test'
)

result_sheet_path=r"C:\Users\patil\OneDrive\Documents\Demo_sheet.xlsx"

# def write_table_in_excel(columns,op,starting_cell,sheet,insert=False,color_fill='FFFF00'):
    
#     print(starting_cell)

#     for n,col in enumerate(columns):
#         curr_row = starting_cell[1]
#         sheet[starting_cell[0]+str(curr_row)]=col
#         sheet[starting_cell[0]+str(curr_row)].border=border
#         sheet[starting_cell[0]+str(curr_row)].fill=fill_color(color_fill)
#         curr_row+=1
#         if n==0 and insert:
#             sheet.insert_rows(curr_row)
        
#         for val in op[col]:
#             sheet[starting_cell[0]+str(curr_row)]=val
#             sheet[starting_cell[0]+str(curr_row)].border=border
#             curr_row+=1
#             if n==0 and insert:
#                 sheet.insert_rows(curr_row)

#         starting_cell[0]=chr(ord(starting_cell[0])+1)

conn = connection.cursor()

border = Border(top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin'))

def fill_color(color) :
    fill = PatternFill(start_color=color,
                   fill_type='solid')
    
    return fill

# Load the workbook
wb = openpyxl.load_workbook(result_sheet_path)

## write values in cell
# # Select the worksheet
sheet = wb['Sheet2']  # Replace 'Sheet1' with the name of your sheet.

# # Write to a specific cell
# sheet['d8'] = 'Hello, New World!'  # Write 'Hello, World!' to cell A1

## to write values of list in same column by adding new rows
# values = ['Value1', 'Value2', 'Value3', 'Value4', 'Value5']

# column = 'D'  # Column D
# start_row = 5

# for value in values:
#     # Find the first empty cell in the specified column
#     cell = sheet[column + str(start_row)]
#     sheet.insert_rows(start_row)
#     sheet[column + str(start_row)] = value
#     start_row += 1


# write function in a cell

# sheet['c5']=10
# sheet['c6']=10
# sheet['c8']='=exact(c5,c6)'

table_name = "hiten.test.dup_test"



query = f"select column_name,data_type,coalesce(character_maximum_length,numeric_precision) character_maximum_length from {table_name.split('.')[0]}.information_schema.columns where lower(table_schema)='{table_name.split('.')[1].lower()}' and lower(table_name)='{table_name.split('.')[2].lower()}' order by ordinal_position"


conn.execute(query)

op = conn.fetch_pandas_all()

columns = op.columns.to_list()


# write_table_in_excel(columns,op,starting_cell=['d',5],sheet=sheet,insert=True,color_fill='40bce6')

write_records_to_excel(op,worksheet=wb['Sheet2_new'],start_cell='d5',insert=True,color_fill='40bce6')

table_name = "hiten.test.dup_test_clone"

query = f"select column_name,data_type,coalesce(character_maximum_length,numeric_precision) character_maximum_length from {table_name.split('.')[0]}.information_schema.columns where lower(table_schema)='{table_name.split('.')[1].lower()}' and lower(table_name)='{table_name.split('.')[2].lower()}' order by ordinal_position"


conn.execute(query)

op = conn.fetch_pandas_all()

# write_table_in_excel(op.columns.to_list(),op,starting_cell=['h',5],sheet=sheet,color_fill='e640d5')

write_records_to_excel(op,worksheet=wb['Sheet2_new'],start_cell='h5',color_fill='e640d5')

# print(op.shape)

starting_cell=['l',5]
src_col,tgt_col = 'd','h'

for cnt,col in enumerate(columns):
    curr_row = starting_cell[1]

    sheet[starting_cell[0]+str(curr_row)]=col
    sheet[starting_cell[0]+str(curr_row)].border=border
    sheet[starting_cell[0]+str(curr_row)].fill=fill_color('40e64e')

    curr_row+=1

    for i in range(curr_row,curr_row+op.shape[0]):
        formula = f"=exact({src_col+str(i)},{tgt_col+str(i)})" if cnt!=len(columns)-1 else f"={src_col+str(i)}>={tgt_col+str(i)}"
        sheet[starting_cell[0]+str(i)]=formula
        sheet[starting_cell[0]+str(i)].border=border

    starting_cell[0]=chr(ord(starting_cell[0])+1)
    src_col=chr(ord(src_col)+1)
    tgt_col=chr(ord(tgt_col)+1)

# Save the workbook
wb.save(result_sheet_path)
