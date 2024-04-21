import openpyxl # pip install openpyxl
import snowflake.connector # pip install snowflake-connector-python
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openpyxl.styles import Border, Side , PatternFill
from openpyxl.utils import get_column_letter

border_style = Border(left=Side(border_style='thin'),
                      right=Side(border_style='thin'),
                      top=Side(border_style='thin'),
                      bottom=Side(border_style='thin'))


def write_records_to_excel(data,worksheet,start_cell,insert=False,color_fill='FFFF00'):
    start_cell=start_cell.upper()

    # Parse row and column indices from the start cell
    match = re.match(r'([A-Z]+)(\d+)', start_cell)
    start_row = int(match.group(2))
    start_col = ord(match.group(1)) - ord('A') + 1

    # Convert the DataFrame to rows
    rows = list(dataframe_to_rows(data, index=False, header=False))
    if insert :
        num_rows_to_insert = len(rows)+3
        worksheet.insert_rows(start_row, amount=num_rows_to_insert)

    # Print column names
    for col_idx, col_name in enumerate(data.columns, start_col):
        cell = worksheet.cell(row=start_row, column=col_idx, value=col_name)
        cell.border = border_style
        cell.fill = fill_color(color_fill)

    # Paste the DataFrame into the selected worksheet starting at the specified cell
    for r_idx, row in enumerate(rows, start_row + 1):  # Start from next row
        for c_idx, value in enumerate(row, start_col):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style




connection = snowflake.connector.connect(
    user='patilhitendra13',
    # authenticator = 'externalbrowser',
    password='Patilhitendra13',
    account='kf94287.ap-south-1.aws',
    warehouse='COMPUTE_WH',
    database='Hiten',
    schema='test'
)

conn = connection.cursor()

result_sheet_path=r"C:\Users\patil\OneDrive\Documents\Demo_sheet.xlsx"


########## filter rows based on  conditions

# table_name = "hiten.test.dup_test"
# query = f"select column_name,data_type,coalesce(character_maximum_length,numeric_precision) character_maximum_length from {table_name.split('.')[0]}.information_schema.columns where lower(table_schema)='{table_name.split('.')[1].lower()}' and lower(table_name)='{table_name.split('.')[2].lower()}' order by ordinal_position"

# conn.execute(query)

# structure = conn.fetch_pandas_all()

# print(structure)

# text_cols = structure[structure['DATA_TYPE']=='TEXT']['COLUMN_NAME'].to_list()
# num_cols = structure[structure['DATA_TYPE']=='NUMBER']['COLUMN_NAME'].to_list()
# print(text_cols,num_cols)


# write table data into excel



def fill_color(color) :
    fill = PatternFill(start_color=color,
                   fill_type='solid')
    
    return fill

query = "select CC_CALL_CENTER_SK,CC_CALL_CENTER_ID,CC_CITY,CC_CLASS,cc_company from SNOWFLAKE_SAMPLE_DATA.TPCDS_SF100TCL.CALL_CENTER limit 10;"
conn.execute(query)
data = conn.fetch_pandas_all()


wb = openpyxl.load_workbook(result_sheet_path)
sheet = wb['Sheet4']

# Select the worksheet you want to paste the DataFrame into
# ws = wb['Sheet3']  # Or you can specify the worksheet by name: wb['Sheet1']

# Define the starting cell where you want to paste the DataFrame
# start_cell = 'c8'

write_records_to_excel(data,worksheet=wb['Sheet4'],start_cell='c8',insert=True,color_fill='40bce6')

new_start_cell = data.shape[0]+8
# print(new_start_cell)
write_records_to_excel(data,worksheet=wb['Sheet4'],start_cell='c'+str(new_start_cell+4),insert=True,color_fill='FF0000')

## comparison result

new_start_cell += (data.shape[0]+5)

num_rows_to_insert = data.shape[0]+5
wb['Sheet4'].insert_rows(new_start_cell, amount=num_rows_to_insert)

# wb['Sheet4']['C'+str(new_start_cell+3)] = 'Start Here'

for curr_col,col in enumerate(data.columns.to_list(),3) : # 3 in enumerate means starting from column C in excel sheet

    curr_row = new_start_cell + 3
    # cell = sheet.cell(row=curr_row, column=curr_col, value=col)
    # cell.border=border_style
    # cell.fill = fill_color('00FFFF')
    # curr_row+=1

    for i in range(curr_row,curr_row+data.shape[0]):
        formula = f"=EXACT({get_column_letter(curr_col)+str((i-(data.shape[0]+4)*2))},{get_column_letter(curr_col)+str((i-(data.shape[0]+4)*1))})"
        cell = sheet.cell(row=i, column=curr_col, value=formula)
        if i==curr_row:
            cell.fill = fill_color('00FF00')
        cell.border=border_style

# Save the workbook
wb.save(result_sheet_path)
